"""
PubMed → Excel  |  Dheeraj Songara
Searches PubMed for each keyword and exports a formatted .xlsx file per query.
"""

# ── Requirements ────────────────────────────────────────────────────────────
print("=" * 55)
print("  PubMed → Excel  |  Required packages to run this:")
print("=" * 55)
print("  pip install requests pandas openpyxl")
print("=" * 55)
print()

# ── Imports ─────────────────────────────────────────────────────────────────
import os
import time
import requests
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── User Input ───────────────────────────────────────────────────────────────
raw = input("Enter keywords (comma-separated):\n> ").strip()
QUERIES = [q.strip() for q in raw.split(",") if q.strip()]

if not QUERIES:
    print("No keywords entered. Exiting.")
    exit()

raw_max = input("\nMax results per query? (press Enter for 60): ").strip()
MAX_RESULTS = int(raw_max) if raw_max.isdigit() else 60

OUTPUT_DIR   = Path(".")
MAX_WORKERS  = 5
NCBI_API_KEY = os.getenv("NCBI_API_KEY", "")

print(f"\n✓ {len(QUERIES)} keyword(s) | max {MAX_RESULTS} results each")
if NCBI_API_KEY:
    print("✓ NCBI API key detected — using 10 req/s")
else:
    print("  No NCBI_API_KEY env var set — using 3 req/s (slower)")
print()

# ── API helpers ──────────────────────────────────────────────────────────────
BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"

def _params(extra: dict) -> dict:
    p = {"db": "pubmed", "retmode": "xml", **extra}
    if NCBI_API_KEY:
        p["api_key"] = NCBI_API_KEY
    return p


def search_pmids(query: str) -> list:
    r = requests.get(
        f"{BASE_URL}/esearch.fcgi",
        params=_params({"term": query, "retmax": MAX_RESULTS}),
        timeout=30,
    )
    r.raise_for_status()
    root = ET.fromstring(r.content)
    return [el.text for el in root.findall(".//Id")]


def fetch_paper(pmid: str) -> dict:
    r = requests.get(
        f"{BASE_URL}/efetch.fcgi",
        params=_params({"id": pmid, "rettype": "abstract"}),
        timeout=30,
    )
    r.raise_for_status()

    root    = ET.fromstring(r.content)
    article = root.find(".//PubmedArticle")
    if article is None:
        return {"pmid": pmid}

    def text(path):
        el = article.find(path)
        return el.text.strip() if el is not None and el.text else ""

    title = text(".//ArticleTitle")

    abstract_els = article.findall(".//AbstractText")
    parts = []
    for ab in abstract_els:
        label = ab.get("Label")
        t = (ab.text or "").strip()
        parts.append(f"{label}: {t}" if label else t)
    abstract = " ".join(parts)

    authors = []
    for a in article.findall(".//Author"):
        last  = (a.findtext("LastName")  or "").strip()
        first = (a.findtext("ForeName") or "").strip()
        if last:
            authors.append(f"{last} {first}".strip())
    authors_str = "; ".join(authors)

    journal = text(".//Journal/Title")
    year    = text(".//PubDate/Year") or text(".//PubDate/MedlineDate")[:4]
    month   = text(".//PubDate/Month")
    date    = f"{year}/{month}" if month else year

    doi = ""
    for id_el in article.findall(".//ArticleId"):
        if id_el.get("IdType") == "doi":
            doi = id_el.text.strip()
            break

    return {
        "pmid":     pmid,
        "doi":      doi,
        "journal":  journal,
        "date":     date,
        "authors":  authors_str,
        "title":    title,
        "abstract": abstract,
    }


def fetch_all_papers(pmids: list) -> list:
    delay   = 0.1 if NCBI_API_KEY else 0.34
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        for pmid in pmids:
            futures[executor.submit(fetch_paper, pmid)] = pmid
            time.sleep(delay)
        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as e:
                print(f"  ✗ PMID {futures[future]}: {e}")
    return results

# ── Excel styling ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ROW_FILL_ALT = PatternFill("solid", start_color="EBF2FA")
THIN_BORDER  = Border(bottom=Side(style="thin", color="CCCCCC"))

COL_WIDTHS = {
    "pmid": 12, "doi": 32, "journal": 28,
    "date": 10, "authors": 36, "title": 50, "abstract": 70,
}


def style_sheet(ws):
    columns = list(COL_WIDTHS.keys())
    for col_idx, col_name in enumerate(columns, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.value     = col_name.upper()
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, ws.max_row + 1):
        fill = ROW_FILL_ALT if row_idx % 2 == 0 else None
        for col_idx in range(1, len(columns) + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == len(columns)))
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_xlsx(records: list, query: str) -> Path:
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in query).strip()
    path      = OUTPUT_DIR / f"{safe_name}.xlsx"
    df = pd.DataFrame(records, columns=["pmid", "doi", "journal", "date", "authors", "title", "abstract"])
    df.sort_values("date", ascending=False, inplace=True)
    df.to_excel(path, index=False, sheet_name="Papers")
    wb = load_workbook(path)
    style_sheet(wb.active)
    wb.save(path)
    return path

# ── Run ──────────────────────────────────────────────────────────────────────
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

for query in QUERIES:
    print(f"🔍 Searching: '{query}'")
    try:
        pmids = search_pmids(query)
    except Exception as e:
        print(f"  ✗ Search failed: {e}\n")
        continue
    if not pmids:
        print("  No results found.\n")
        continue
    print(f"  Found {len(pmids)} papers. Fetching details…")
    records = fetch_all_papers(pmids)
    path    = save_xlsx(records, query)
    print(f"  ✓ Saved {len(records)} papers → {path.name}\n")

print("Done! All .xlsx files saved in the current folder.")
